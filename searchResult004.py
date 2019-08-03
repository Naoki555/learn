#coding: UTF-8
import urllib.request
import xml.etree.ElementTree as ET
from openpyxl import Workbook
import openpyxl
import datetime
import sys
#------------------------------
# Excelのシート初期化メソッド
#------------------------------
def initializeSheet(ws):
  
  #タイトル設定
  ws['B2']="日付"
  ws['C2']="検索人数"
  ws['D2']="順位"
  ws['E2']="プラン名"
  ws['F2']="2食付き最安価格/1名単価"
  ws['G2']="ホテル名"
  ws['H2']="住所"
  ws['I2']="タイプ"
  ws['J2']="5位のプラン"
  ws['K2']="5位のホテル名"
  ws['L2']="5位以内に入る単価"

  #セルの幅調整
  ws.column_dimensions['A'].width = 5
  ws.column_dimensions['B'].width = 15
  ws.column_dimensions['C'].width = 10
  ws.column_dimensions['D'].width = 5
  ws.column_dimensions['E'].width = 30
  ws.column_dimensions['F'].width = 15
  ws.column_dimensions['G'].width = 30
  ws.column_dimensions['H'].width = 30
  ws.column_dimensions['I'].width = 10
  ws.column_dimensions['J'].width = 20
  ws.column_dimensions['K'].width = 20
  ws.column_dimensions['L'].width = 20

  #タイトルの色設定
  fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='90EE90')
  ws['A2'].fill = fill
  ws['B2'].fill = fill
  ws['C2'].fill = fill
  ws['D2'].fill = fill
  ws['E2'].fill = fill
  ws['F2'].fill = fill
  ws['G2'].fill = fill
  ws['H2'].fill = fill
  ws['I2'].fill = fill
  ws['J2'].fill = fill
  ws['K2'].fill = fill
  ws['L2'].fill = fill
#------------------------------
# Excelのシートに出力するメソッド
#------------------------------
def outputOnSheet(ws, target_date, plan_list, index, row):
  print('[Target date]:' + str(target_date))
  #プランリストの中から5番目（1位）のランク、プラン名、プランの価格、ホテル名、価格、住所、ホテルタイプを出力
  ws['B' + str(row)] = target_date.strftime("%m/%d (%a)")
  ws['C' + str(row)] = index
  if len(plan_list) >= 1 :  #配列の長さが5以上
    ws['D' + str(row)] = plan_list[0]['rank']
    ws['E' + str(row)] = plan_list[0]['plan_name']
    ws['F' + str(row)].number_format = "#,##0" 
    ws['F' + str(row)] = plan_list[0]['plan_value']
    ws['G' + str(row)] = plan_list[0]['hotel_name']
    ws['H' + str(row)] = plan_list[0]['hotel_address']
    ws['I' + str(row)] = plan_list[0]['hotel_type']
  if len(plan_list) >= 5 :  #配列の長さが5以上  
    #プランリストの中から5番目（5位）のプラン名、ホテル名、価格を出力
    ws['J' + str(row)] = plan_list[4]['plan_name']
    ws['K' + str(row)] = plan_list[4]['hotel_name']
    ws['L' + str(row)].number_format = "#,##0" 
    ws['L' + str(row)] = plan_list[4]['plan_value']
  row = row + 1

#------------------------------
# API経由で情報を取得するメソッド
#------------------------------
def apicall(s_area ,_stay_date, adult_num):
  key = "peg16bf86efe77" # じゃらんで取得したAPIキー
  #pref = ""         # 都道府県指定：       大分県 440000
  #l_area = ""       # 大エリアコード指定： 湯布院 440600
  #s_area = "440602" # 小エリアコード指定： 湯布院・湯平 440602
  #stay_date ='20190901' #宿泊日を指定する
  stay_date = str(_stay_date.year) + str(_stay_date.month) + str(_stay_date.day)
  stay_date = _stay_date.strftime('%Y%m%d')
  print(stay_date)
  stay_count = 1  # 宿泊数
  room_count = 1   # 部屋数

  # adult_num = "1"   # 一部屋あたりの大人の数2名から6名
  h_type = 4      # 絞り込みなし0 1:旅館 2:ペンション・民宿・ロッジ 3:貸し別荘・コンドミニアム 4:ホテル・ビジネスホテル 5:公共の宿
  _2_meals = 1     # 朝夕食事アリの宿を絞る
  order = 2        # 安い順：２
  count = 5        # 表示件数
  xml_ptn = 2       # 基本情報のみ0 プラン含む:2


  url = 'http://jws.jalan.net/APIAdvance/StockSearch/V1/'
  params = {'key': key, 's_area': s_area, 'count': count, 'xml_ptn':xml_ptn ,  '2_meals' : _2_meals , 'order' : order , 'h_type' : h_type, 'stay_count' :stay_count , 'room_count' : room_count, 'adult_num' :adult_num, 'stay_date' : stay_date }
  OutPlanList = []

  req = urllib.request.Request('{}?{}'.format(url, urllib.parse.urlencode(params)))
  with urllib.request.urlopen(req) as response:
    XmlData = response.read()
    root = ET.fromstring(XmlData)
    # print(root.tag,root.text)
    rank = 0  #順位
    # 子階層のタグと中身
    for planList in root:
        print(planList.tag, planList.text)
        if planList.tag == '{jws}Plan':
          rank = rank + 1
          _plan_name = ''
          _plan_value =''
          _hotel_name =''
          _hotel_address = ''
          _hotel_type =''
          # Hotelの基本情報
          for plan in planList:
            print(plan.tag, plan.text)
            if plan.tag == '{jws}PlanName':
              _plan_name = plan.text
            if plan.tag == '{jws}SampleRate':
              _plan_value = plan.text
            if plan.tag == '{jws}Hotel':
              for hotel in plan:
                if hotel.tag == '{jws}HotelName':
                  print(hotel.tag, hotel.text)
                  _hotel_name = hotel.text
                if hotel.tag == '{jws}HotelAddress':
                  print(hotel.tag, hotel.text)
                  _hotel_address = hotel.text
                if hotel.tag == '{jws}HotelType':
                  print(hotel.tag, hotel.text)
                  _hotel_type = hotel.text
              item ={
                'rank': rank,
                'plan_name': _plan_name,
                'plan_value':_plan_value,
                'hotel_name' : _hotel_name,
                'plan_value':int(_plan_value)/110*100,
                'hotel_address' :_hotel_address,
                'hotel_type': _hotel_type
              }
              OutPlanList.append(item)
  return OutPlanList
#------------------------------
# Main 処理
#------------------------------
if __name__ == "__main__":

  row = 3 #３行目からデータ出力開始
  if len(sys.argv) != 3: #引数にエリアコードと日付を指定する,なければエラーを表示して終了
    print('+--------------------------------------------------------------+')
    print(' 注意：引数に、検索対象の小エリアコードと、検索日を指定してください')
    print(' 例：小エリアコード(熱海)： 210202  検索日：2019年9月1日')
    print(' >py searchResult004.py 210202 20191001')
    print('+--------------------------------------------------------------+')
    exit()

  s_area = str(sys.argv[1]) # '440602' 第一引数は小エリアコード
  target_date = str(sys.argv[2]) # '20190901'第二引数は検索対象の日付の1日目
  
  year = int(target_date[0:4])
  month = int(target_date[4:6])
  day = int(target_date[6:8])
  target_date = datetime.date(year, month, day)#ストリング型から日付型に変換する
  wb = Workbook()
  ws = wb.active
  ws.title = "安いホテルリスト"
  #シートの初期化
  initializeSheet(ws)

  for term in range(0,10): #7日分のデータを出力
    for i in range(1, 7): #6人までの単価を検索する
      # API経由で情報取得
      plan_list = apicall(s_area, target_date, str(i)) # 何名様かをパラメータで渡す
      # Excelのシート作成
      outputOnSheet(ws,target_date,plan_list, i, row + (i - 1) + term * 10)
    target_date = target_date + datetime.timedelta(days=1) #検索対象日をプラス1日して、再度検索を行う

  # Excel出力
  file_name = "じゃらん安いホテルリスト" + datetime.datetime.today().strftime("%Y%m%d_%H%M%S") + '.xlsx'
  wb.save(file_name)

